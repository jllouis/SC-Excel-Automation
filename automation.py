# Copyright (c) 2018, Lenovo
# Author: Jose L. Louis, <jlouis@lenovo.com>

import openpyxl  # needed to make Python interact with excel docs
from openpyxl.utils import get_column_letter, column_index_from_string


# from optparse import OptionParser  # used to get filename and output file name from command line arguments (-f and -o)


# def parse_opts():
#     # creating and configuring command line options parser
#     parser = OptionParser(
#         description="Processes Global Supply data sheets and outputs summary workbook of items at risk of imminent "
#                     "depletion.",
#         version="%prog 1.0")
#
#     # get the (-f) input file and (-o) desired name of the output file
#     parser.add_option("-f", "--file", dest="filename", metavar="INPUT_FILE_NAME",
#                       help="The input file to process.")
#     parser.add_option("-o", "--outfile", dest="output", metavar="OUTPUT_FILE_NAME",
#                       help="The name of the output file.")
#
#     (options, args) = parser.parse_args()
#
#     # in case of no arguments
#     if len(args) == 0:
#         parser.print_help()
#
#     # run the main process function with the supplied file names
#     process(options.filename, options.output)


# filters file based on given criteria in our case "chassis" and "bdplannar"
def filter_data(data, criteria):
    fits_criteria = []
    headers = []
    # filter based on criteria and DOI amount
    for i, datum in enumerate(data[3]):
        if i == 0:
            for j in range(0, 16):
                headers.append(data[j][i])
            continue
        if datum.value in criteria:
            tmp = []
            if data[7][i].value < 25:  # include data if only DOI is less than 25
                for j in range(0, 16):
                    tmp.append(data[j][i])
                fits_criteria.append(tmp)

    return fits_criteria


# write result to excel document
def export_data(data, output):
    out_wb = openpyxl.Workbook()
    out_sheet = out_wb.active
    out_sheet.title = 'Result'

    # write header to excel file
    # for i, words in enumerate(header):
    #     for j, value in enumerate(words):
    #         out_sheet.cell(row=1 + i, column=1 + j).value = value.value

    # write to data excel file

    # re-order row
    re_ordered = [
        [row[0], row[1], row[2], row[10], row[5], row[6], row[7], row[4], row[11], row[12], row[13], row[14], row[15]]
        for row in data]
    for i, datum in enumerate(re_ordered):
        for j, value in enumerate(datum):
            out_sheet.cell(row=2 + i, column=1 + j).value = value.value

    out_wb.save(filename=output)


def process(filename, output):
    print('Reading data...')
    print("filename is " + filename)
    wb = openpyxl.load_workbook(filename)
    print('Reading Done!')

    # prepare data
    global_supply = wb.get_sheet_by_name('GLOBAL_SUPPLY')
    columns = ['A', 'B', 'C', 'D', 'AT', 'AW', 'AX', 'BC', 'DH', 'DI', 'DJ', 'DO', 'DP', 'DQ', 'DR', 'DW']
    column_data = list(global_supply.columns)
    filtered_data = list(list(column_data[column_index_from_string(x) - 1]) for x in columns)

    filter_result = filter_data(filtered_data, ['CHASSIS', 'BDPLANAR'])

    # export to excel
    export_data(filter_result, output)

    print('Finished.')


def main():
    # import os
    # import sys
    # results = []
    # for file in os.listdir(sys.path[0]):
    #     print("found ", file)
    #     if file.endswith(".xlsx"):
    #         results.append(file)
    #
    # if len(results) > 1:
    #     print("Error, more than 1 .xlsx file in directory!")
    #     exit(-1)
    file = input("Enter file path: ")
    output = input("Enter output file name: ")
    print("processing ", file)
    process(file.strip("\""), output)
    print("Processing Done!")


if __name__ == "__main__":
    main()

# parse_opts()
# # print('ok')
